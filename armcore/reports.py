"""
Два Excel-реестра по итогам работы.

ТЗ, п.9:
  1) Первый Excel — только иностранные граждане, проходящие обучение
     (на них формируется договор и акт на обучение).
  2) Второй Excel — все остальные.

Столбцы (ТЗ): ФИО, дата рождения, номер паспорта, дата выдачи и орган,
гражданство, дата обращения, оказанные услуги.

openpyxl импортируется лениво — чтобы модуль импортировался даже там, где
openpyxl/numpy окружения временно сломаны (на целевой Windows-машине работает).
"""
from __future__ import annotations

import os
from datetime import datetime
from typing import Dict, List

from . import services as services_mod


HEADERS = [
    "ФИО",
    "Дата рождения",
    "Номер паспорта",
    "Дата выдачи и орган",
    "Гражданство",
    "Дата обращения",
    "Оказанные услуги",
    "Комментарий",
]


def _row_from_fields(fields: Dict[str, str], selected: List[str],
                     comment: str = "") -> List[str]:
    date_issue = fields.get("DATE_ISSUE", "")
    issued_by = fields.get("ISSUED_BY", "")
    date_place = " ".join(p for p in (date_issue, issued_by) if p).strip()
    return [
        fields.get("FIO", ""),
        fields.get("BIRTHDAY", ""),
        fields.get("PASSPORT_NUMBER", ""),
        date_place,
        fields.get("COUNTRY_CODE", ""),
        datetime.now().strftime("%d.%m.%Y"),
        ", ".join(selected),
        comment,
    ]


def _person_exists(ws, fio: str, passport: str) -> bool:
    """Есть ли уже запись с тем же ФИО и номером паспорта (защита от дублей)."""
    fio_l = (fio or "").strip().lower()
    passport_s = (passport or "").strip()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        r_fio = (row[0] or "").strip().lower() if len(row) > 0 else ""
        r_pass = (str(row[2]).strip() if len(row) > 2 and row[2] else "")
        if r_fio == fio_l and r_pass == passport_s and passport_s:
            return True
    return False


def append_record(excel_path: str, fields: Dict[str, str],
                  selected: List[str], comment: str = "",
                  skip_duplicates: bool = True) -> bool:
    """Добавить строку в указанный Excel-реестр (создаёт файл при необходимости).

    Возвращает True, если строка добавлена; False, если пропущена как дубль.
    """
    from openpyxl import Workbook, load_workbook  # ленивый импорт

    os.makedirs(os.path.dirname(os.path.abspath(excel_path)), exist_ok=True)

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Реестр"
        ws.append(HEADERS)

    fio = fields.get("FIO", "")
    passport = fields.get("PASSPORT_NUMBER", "")
    if skip_duplicates and _person_exists(ws, fio, passport):
        return False

    ws.append(_row_from_fields(fields, selected, comment))
    wb.save(excel_path)
    return True


def record_client(cfg, fields: Dict[str, str], selected: List[str],
                  comment: str = "") -> str:
    """Записать клиента в нужный реестр по правилу ТЗ (обучение / прочие).

    cfg — armcore.config.Config (нужны report_study_path / report_other_path).
    Возвращает путь к реестру, в который попала запись.
    """
    if services_mod.is_study_client(selected):
        path = cfg.report_study_path
    else:
        path = cfg.report_other_path
    append_record(path, fields, selected, comment=comment)
    return path
